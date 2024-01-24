#!/usr/bin/python3

import sys

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


def get_answer_distribution_for_question(dataset):
    question = dataset[0]
    answers = dataset[1:]

    distribution = {}

    for answer in answers:
        if answer == None:
            continue

        if answer not in distribution:
            distribution[answer] = 0
        distribution[answer] += 1

    return (question, distribution)


def get_answer_distribution_for_multiple_options_question(*args):
    datasets = list(*args)

    headers = [dataset[0] for dataset in datasets]
    question = get_part_before_hyphen(headers[0])
    options = [option[len(question + ' - '):] for option in headers]
    responses_for_options = [column[1:] for column in datasets]

    distribution = {option: {} for option in options}

    for option_id, responses_for_option in enumerate(responses_for_options):
        option = options[option_id]

        for response in responses_for_option:
            if response == '' or response == None:
                continue

            if response not in distribution[option]:
                distribution[option][response] = 0
            distribution[option][response] += 1

    if all([distribution_for_option.keys() == [0, 1] for distribution_for_option in distribution.values()]):
        aggregated_distribution = {option: values[1]
                                   for option, values in distribution.items()}
        return [(question, aggregated_distribution)]

    return [(question, distribution)]


def get_part_before_hyphen(string):
    return string.split(' - ')[0]


def group_related_columns(columns):
    headers = [column_id for column_id, _ in enumerate(columns)]
    headers.sort(key=lambda x: columns[x][0])
    groups = []

    start_of_group = 0
    while start_of_group < len(headers):
        prefix = get_part_before_hyphen(columns[headers[start_of_group]][0])
        end_of_group = start_of_group + 1
        while end_of_group < len(headers) and get_part_before_hyphen(columns[headers[end_of_group]][0]) == prefix:
            end_of_group += 1

        group = headers[start_of_group:end_of_group]
        if len(group) > 1:
            column_group = list(map(lambda x: columns[x], group))
            groups.append(column_group)
        else:
            groups.append(columns[group[0]])

        start_of_group = end_of_group

    return groups


def write_distribution_to_sheet_for_multiple_options_question(distsheet, question, distribution):
    distsheet.append((question,))
    distsheet.append(tuple([''] + list(list(distribution.values())[0].keys())))
    for option, answers in distribution.items():
        distsheet.append(tuple([option] + list(answers.values())))


def write_distribution_to_sheet_for_single_option_question(distsheet, question, distribution):
    distsheet.append((question,))
    for answer, count in sorted(distribution.items(), key=lambda kv : kv[0]):
        distsheet.append((answer, count))


def write_distribution_to_sheet(distsheet, question, distribution):
    if type(list(distribution.values())[0]) == dict:
        return write_distribution_to_sheet_for_multiple_options_question(
            distsheet, question, distribution)
    else:
        return write_distribution_to_sheet_for_single_option_question(
            distsheet, question, distribution)


def draw_distribution_chart_for_multiple_options_question(distsheet, row_idx, chartsheet, question, distribution):
    title_row = row_idx  # solely for debug purposes
    option_count = max([len(distribution_for_option.items())
                       for distribution_for_option in distribution.values()])
    entry_count = len(distribution.items())
    header_row = row_idx + 1
    first_data_row = row_idx + 1 + 1
    last_data_row = row_idx + 1 + entry_count

    chart = BarChart()
    data = Reference(distsheet, min_col=2, max_col=option_count + 1,
                       min_row=header_row, max_row=last_data_row)
    labels = Reference(distsheet, min_col=1, max_col=1,
                     min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = question
    chartsheet.add_chart(chart)

    return entry_count + 2


def draw_distribution_chart_for_single_option_question(distsheet, row_idx, chartsheet, question, distribution):
    title_row = row_idx  # solely for debug purposes
    entry_count = len(distribution.items())
    first_data_row = row_idx + 1
    last_data_row = row_idx + entry_count

    chart = PieChart()
    labels = Reference(distsheet, min_col=1,
                       min_row=first_data_row, max_row=last_data_row)
    data = Reference(distsheet, min_col=2,
                     min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.title = question
    chartsheet.add_chart(chart)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.dLblPos = 'outEnd'
    chart.dataLabels.showPercent = True

    return entry_count + 1


def draw_distribution_chart(distsheet, row_idx, chartsheet, question, distribution):
    if type(list(distribution.values())[0]) == dict:
        return draw_distribution_chart_for_multiple_options_question(
            distsheet, row_idx, chartsheet, question, distribution)
    else:
        return draw_distribution_chart_for_single_option_question(
            distsheet, row_idx, chartsheet, question, distribution)


def make_charts(inputfile, outputfile):
    workbook = load_workbook(filename=inputfile)

    datasheet = workbook.active

    columns = [column for column in datasheet.iter_cols(values_only=True)]
    column_groups = group_related_columns(columns)

    distributions = []

    for dataset in column_groups:
        if type(dataset) == list:
            distributions += get_answer_distribution_for_multiple_options_question(
                dataset)
        else:
            distributions.append(get_answer_distribution_for_question(dataset))

    distsheet = workbook.create_sheet("distributions")
    for question, distribution in distributions:
        write_distribution_to_sheet(distsheet, question, distribution)

    sheet_idx = 0
    row_idx = 1
    for question, distribution in distributions:
        chartsheet = workbook.create_chartsheet('diagram%d' % sheet_idx)

        entry_count = draw_distribution_chart(
            distsheet, row_idx, chartsheet, question, distribution)

        row_idx += entry_count
        sheet_idx += 1

    workbook.save(outputfile)


if __name__ == '__main__':
    inputfile = 'responses.xlsx'
    if len(sys.argv) > 1:
        inputfile = sys.argv[1]
    outputfile = 'analytics.xlsx'
    if len(sys.argv) > 2:
        outputfile = sys.argv[2]

    make_charts(inputfile, outputfile)
